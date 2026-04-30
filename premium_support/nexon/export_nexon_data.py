import json
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/haewon.yum/Documents/Queries/premiume_support/nexon/nexon_ios_analysis.xlsx"

def run_bq(query):
    result = subprocess.run(
        ["bq", "query", "--use_legacy_sql=false", "--format=json", "--max_rows=500", query],
        capture_output=True, text=True, timeout=180
    )
    if result.returncode != 0:
        raise Exception(f"BQ error: {result.stderr}")
    return json.loads(result.stdout)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_sheet(wb, sheet_name, data, columns, col_labels=None):
    ws = wb.create_sheet(title=sheet_name)
    labels = col_labels or columns
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for r, row in enumerate(data, 2):
        for c, col in enumerate(columns, 1):
            val = row.get(col)
            if val is not None and val != "":
                try:
                    val = float(val)
                    if val == int(val) and abs(val) < 1e15:
                        val = int(val)
                except (ValueError, OverflowError):
                    pass
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    for c in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(labels[c-1]) + 4)

# --- Queries ---

q_daily_spend_os = """
SELECT date_utc, campaign.os AS os,
  ROUND(SUM(gross_spend_usd), 2) AS gross_spend_usd,
  SUM(impressions) AS impressions, SUM(clicks) AS clicks, SUM(installs) AS installs,
  ROUND(SAFE_DIVIDE(SUM(gross_spend_usd), SUM(installs)), 2) AS cpi
FROM `moloco-ae-view.athena.fact_dsp_core`
WHERE campaign.tracking_entity = 'NEXON'
  AND date_utc BETWEEN DATE_SUB(CURRENT_DATE(), INTERVAL 30 DAY) AND DATE_SUB(CURRENT_DATE(), INTERVAL 1 DAY)
GROUP BY 1, 2 ORDER BY 1, 2
"""

q_acs_overall = """
SELECT
  CASE WHEN date_utc < '2026-02-12' THEN 'PRE (Jan22-Feb11)' ELSE 'POST (Feb12-Feb22)' END AS period,
  COUNT(DISTINCT date_utc) AS days,
  ROUND(SUM(gross_spend_usd), 2) AS total_spend,
  ROUND(SUM(gross_spend_usd) / COUNT(DISTINCT date_utc), 2) AS avg_daily_spend,
  SUM(impressions) AS impressions, SUM(clicks) AS network_clicks, SUM(clicks_ev) AS true_clicks,
  SUM(installs) AS installs,
  ROUND(SUM(installs) / COUNT(DISTINCT date_utc), 0) AS avg_daily_installs,
  ROUND(SAFE_DIVIDE(SUM(gross_spend_usd), SUM(installs)), 2) AS cpi,
  ROUND(SAFE_DIVIDE(SUM(clicks), SUM(impressions)) * 100, 3) AS net_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(clicks_ev), SUM(impressions)) * 100, 3) AS true_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks)) * 100, 3) AS net_cvr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks_ev)) * 100, 3) AS true_cvr_pct,
  ROUND(SAFE_DIVIDE(SUM(gross_spend_usd), SUM(impressions)) * 1000, 2) AS cpm
FROM `moloco-ae-view.athena.fact_dsp_creative`
WHERE advertiser.mmp_bundle_id = '6739616715' AND campaign.os = 'IOS'
  AND date_utc BETWEEN '2026-01-22' AND '2026-02-22'
  AND date_utc != '2026-02-19' AND gross_spend_usd > 0
GROUP BY 1 ORDER BY 1
"""

q_acs_campaign = """
SELECT
  CASE WHEN date_utc < '2026-02-12' THEN 'PRE' ELSE 'POST' END AS period,
  campaign.title AS campaign,
  COUNT(DISTINCT date_utc) AS days,
  ROUND(SUM(gross_spend_usd), 2) AS total_spend,
  ROUND(SUM(gross_spend_usd) / COUNT(DISTINCT date_utc), 2) AS avg_daily_spend,
  SUM(installs) AS installs,
  ROUND(SUM(installs) / COUNT(DISTINCT date_utc), 0) AS avg_daily_installs,
  ROUND(SAFE_DIVIDE(SUM(gross_spend_usd), SUM(installs)), 2) AS cpi,
  ROUND(SAFE_DIVIDE(SUM(clicks), SUM(impressions)) * 100, 3) AS net_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks)) * 100, 3) AS net_cvr_pct
FROM `moloco-ae-view.athena.fact_dsp_creative`
WHERE advertiser.mmp_bundle_id = '6739616715' AND campaign.os = 'IOS'
  AND date_utc BETWEEN '2026-01-22' AND '2026-02-22'
  AND date_utc != '2026-02-19' AND gross_spend_usd > 0
GROUP BY 1, 2 HAVING total_spend > 100 ORDER BY 2, 1
"""

q_acs_format = """
SELECT
  CASE WHEN date_utc < '2026-02-12' THEN 'PRE' ELSE 'POST' END AS period,
  creative.format AS cr_format,
  ROUND(SUM(gross_spend_usd), 2) AS total_spend,
  ROUND(SUM(gross_spend_usd) / COUNT(DISTINCT date_utc), 2) AS avg_daily_spend,
  SUM(impressions) AS impressions,
  SUM(clicks) AS network_clicks, SUM(clicks_ev) AS true_clicks,
  SUM(installs) AS installs,
  ROUND(SAFE_DIVIDE(SUM(gross_spend_usd), SUM(installs)), 2) AS cpi,
  ROUND(SAFE_DIVIDE(SUM(clicks), SUM(impressions)) * 100, 3) AS net_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(clicks_ev), SUM(impressions)) * 100, 3) AS true_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks)) * 100, 3) AS net_cvr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks_ev)) * 100, 3) AS true_cvr_pct
FROM `moloco-ae-view.athena.fact_dsp_creative`
WHERE advertiser.mmp_bundle_id = '6739616715' AND campaign.os = 'IOS'
  AND date_utc BETWEEN '2026-01-22' AND '2026-02-22'
  AND date_utc != '2026-02-19' AND gross_spend_usd > 0
GROUP BY 1, 2 ORDER BY 2, 1
"""

q_sor = """
SELECT install_date_utc,
  SUM(CASE WHEN is_attributed THEN installs ELSE 0 END) AS moloco_installs,
  SUM(installs) AS total_installs,
  ROUND(SAFE_DIVIDE(SUM(CASE WHEN is_attributed THEN installs ELSE 0 END), SUM(installs)) * 100, 2) AS soi_pct,
  ROUND(SUM(CASE WHEN is_attributed THEN revenue_usd_d1 ELSE 0 END), 2) AS moloco_rev_d1,
  ROUND(SUM(revenue_usd_d1), 2) AS total_rev_d1,
  ROUND(SAFE_DIVIDE(SUM(CASE WHEN is_attributed THEN revenue_usd_d1 ELSE 0 END), SUM(revenue_usd_d1)) * 100, 2) AS sor_d1_pct,
  ROUND(SUM(CASE WHEN is_attributed THEN revenue_usd_d3 ELSE 0 END), 2) AS moloco_rev_d3,
  ROUND(SUM(revenue_usd_d3), 2) AS total_rev_d3,
  ROUND(SAFE_DIVIDE(SUM(CASE WHEN is_attributed THEN revenue_usd_d3 ELSE 0 END), SUM(revenue_usd_d3)) * 100, 2) AS sor_d3_pct,
  ROUND(SUM(CASE WHEN is_attributed THEN revenue_usd_d7 ELSE 0 END), 2) AS moloco_rev_d7,
  ROUND(SUM(revenue_usd_d7), 2) AS total_rev_d7,
  ROUND(SAFE_DIVIDE(SUM(CASE WHEN is_attributed THEN revenue_usd_d7 ELSE 0 END), SUM(revenue_usd_d7)) * 100, 2) AS sor_d7_pct
FROM `moloco-ae-view.market_share.fact_market_event`
WHERE mmp_bundle_id = '6739616715' AND os = 'IOS'
  AND install_date_utc BETWEEN DATE_SUB(CURRENT_DATE(), INTERVAL 30 DAY) AND DATE_SUB(CURRENT_DATE(), INTERVAL 1 DAY)
GROUP BY 1 ORDER BY 1
"""

q_daily_ios = """
SELECT date_utc,
  ROUND(SUM(gross_spend_usd), 2) AS gross_spend_usd,
  SUM(impressions) AS impressions,
  SUM(clicks) AS network_clicks, SUM(clicks_ev) AS true_clicks,
  SUM(installs) AS installs,
  ROUND(SAFE_DIVIDE(SUM(gross_spend_usd), SUM(installs)), 2) AS cpi,
  ROUND(SAFE_DIVIDE(SUM(clicks), SUM(impressions)) * 100, 3) AS net_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(clicks_ev), SUM(impressions)) * 100, 3) AS true_ctr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks)) * 100, 3) AS net_cvr_pct,
  ROUND(SAFE_DIVIDE(SUM(installs), SUM(clicks_ev)) * 100, 3) AS true_cvr_pct,
  COUNT(DISTINCT creative.format) AS format_count,
  COUNT(DISTINCT creative.id) AS unique_creatives
FROM `moloco-ae-view.athena.fact_dsp_creative`
WHERE advertiser.mmp_bundle_id = '6739616715' AND campaign.os = 'IOS'
  AND date_utc BETWEEN DATE_SUB(CURRENT_DATE(), INTERVAL 60 DAY) AND DATE_SUB(CURRENT_DATE(), INTERVAL 1 DAY)
  AND gross_spend_usd > 0
GROUP BY 1 ORDER BY 1
"""

print("Running queries...")

print("  1/6 Daily spend by OS...")
d1 = run_bq(q_daily_spend_os)
print("  2/6 ACS overall...")
d2 = run_bq(q_acs_overall)
print("  3/6 ACS by campaign...")
d3 = run_bq(q_acs_campaign)
print("  4/6 ACS by format...")
d4 = run_bq(q_acs_format)
print("  5/6 SOR trend...")
d5 = run_bq(q_sor)
print("  6/6 Daily iOS performance...")
d6 = run_bq(q_daily_ios)

print("Writing Excel...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

write_sheet(wb, "Daily Spend by OS", d1,
    ["date_utc", "os", "gross_spend_usd", "impressions", "clicks", "installs", "cpi"],
    ["Date", "OS", "Spend ($)", "Impressions", "Clicks", "Installs", "CPI ($)"])

write_sheet(wb, "ACS Pre-Post Overall", d2,
    ["period", "days", "total_spend", "avg_daily_spend", "impressions", "network_clicks", "true_clicks",
     "installs", "avg_daily_installs", "cpi", "net_ctr_pct", "true_ctr_pct", "net_cvr_pct", "true_cvr_pct", "cpm"],
    ["Period", "Days", "Total Spend ($)", "Avg Daily Spend ($)", "Impressions", "Network Clicks", "True Clicks",
     "Installs", "Avg Daily Installs", "CPI ($)", "Net CTR%", "True CTR%", "Net CVR%", "True CVR%", "CPM ($)"])

write_sheet(wb, "ACS Pre-Post Campaign", d3,
    ["period", "campaign", "days", "total_spend", "avg_daily_spend", "installs", "avg_daily_installs", "cpi", "net_ctr_pct", "net_cvr_pct"],
    ["Period", "Campaign", "Days", "Total Spend ($)", "Avg Daily Spend ($)", "Installs", "Avg Daily Installs", "CPI ($)", "Net CTR%", "Net CVR%"])

write_sheet(wb, "ACS Pre-Post Format", d4,
    ["period", "cr_format", "total_spend", "avg_daily_spend", "impressions", "network_clicks", "true_clicks",
     "installs", "cpi", "net_ctr_pct", "true_ctr_pct", "net_cvr_pct", "true_cvr_pct"],
    ["Period", "Format", "Total Spend ($)", "Avg Daily Spend ($)", "Impressions", "Network Clicks", "True Clicks",
     "Installs", "CPI ($)", "Net CTR%", "True CTR%", "Net CVR%", "True CVR%"])

write_sheet(wb, "SOR-SOI Trend", d5,
    ["install_date_utc", "moloco_installs", "total_installs", "soi_pct",
     "moloco_rev_d1", "total_rev_d1", "sor_d1_pct",
     "moloco_rev_d3", "total_rev_d3", "sor_d3_pct",
     "moloco_rev_d7", "total_rev_d7", "sor_d7_pct"],
    ["Date", "Moloco Installs", "Total Installs", "SOI%",
     "Moloco Rev d1 ($)", "Total Rev d1 ($)", "SOR d1%",
     "Moloco Rev d3 ($)", "Total Rev d3 ($)", "SOR d3%",
     "Moloco Rev d7 ($)", "Total Rev d7 ($)", "SOR d7%"])

write_sheet(wb, "Daily iOS Performance", d6,
    ["date_utc", "gross_spend_usd", "impressions", "network_clicks", "true_clicks",
     "installs", "cpi", "net_ctr_pct", "true_ctr_pct", "net_cvr_pct", "true_cvr_pct",
     "format_count", "unique_creatives"],
    ["Date", "Spend ($)", "Impressions", "Network Clicks", "True Clicks",
     "Installs", "CPI ($)", "Net CTR%", "True CTR%", "Net CVR%", "True CVR%",
     "Format Count", "Unique Creatives"])

wb.save(OUTPUT_PATH)
print(f"Done! Saved to {OUTPUT_PATH}")
